import openpyxl
import torch


class AverageMeter(object):
    """
    Computes and stores the average and
    current value.
    """
    def __init__(self):
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count

def accuracy(output, target):
    preds = output.argmax(dim=1)
    correct_k = torch.sum(preds.view(-1) == target.view(-1)).item()

    return correct_k


def getWorkBook():
    mywb = openpyxl.Workbook()
    mywb.create_sheet(index=0,title='epoch_trainloss')
    mywb.create_sheet(index=1,title='epoch_trainacc')
    mywb.create_sheet(index=2,title='epoch_testloss')
    mywb.create_sheet(index=3,title='epoch_testacc')
    return mywb